import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime
import PyPDF2
import pdfplumber
from typing import Dict, List, Optional, Tuple
import logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configuration de la page
st.set_page_config(
    page_title="PDF Extractor Select T.T",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Configuration du logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class PDFExtractor:
    """Classe pour extraire les données des auto-factures Select T.T"""
    
    def __init__(self):
        self.extracted_data = []
    
    def normalize_amount(self, amount_str: str) -> Optional[float]:
        """Normalise les montants avec différents formats"""
        try:
            if not amount_str:
                return None
            
            amount_str = str(amount_str).strip()
            
            if ',' in amount_str and '.' in amount_str:
                amount_str = amount_str.replace(',', '')
            elif ',' in amount_str:
                amount_str = amount_str.replace(',', '.')
            
            amount_str = re.sub(r'[^\d\.]', '', amount_str)
            
            result = float(amount_str)
            return result if not pd.isna(result) else None
            
        except (ValueError, TypeError):
            logger.warning(f"Impossible de normaliser le montant: {amount_str}")
            return None
    
    def extract_with_regex(self, text: str, pattern: str, group: int = 1) -> Optional[str]:
        """Extrait une valeur avec regex de manière sécurisée"""
        try:
            match = re.search(pattern, text, re.IGNORECASE)
            return match.group(group).strip() if match and match.group(group) else None
        except (AttributeError, IndexError):
            return None
    
    def extract_invoice_data(self, pdf_file) -> Dict:
        """Extrait les données d'une facture PDF"""
        try:
            with pdfplumber.open(pdf_file) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
            
            if not text.strip():
                pdf_file.seek(0)
                reader = PyPDF2.PdfReader(pdf_file)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() or ""
            
            logger.info(f"Texte extrait (premiers 200 chars): {text[:200]}...")
            
            data = {
                'nom_fichier': pdf_file.name,
                'numero_facture': self.extract_invoice_number(text),
                'date_facture': self.extract_invoice_date(text),
                'numero_commande': self.extract_purchase_order(text),
                'date_echeance': self.extract_due_date(text),
                'destinataire': self.extract_recipient(text),
                'emetteur': 'Select T.T',
                'batch_id': self.extract_batch_id(text),
                'assignment_id': self.extract_assignment_id(text),
                'total_net': self.extract_total_net(text),
                'total_tva': self.extract_total_vat(text),
                'total_brut': self.extract_total_gross(text),
                'taux_tva': '20%',
                'devise': 'EUR',
                'lignes_detail': self.extract_line_items(text),
                'rubriques_analyse': None
            }
            
            if data['lignes_detail']:
                data['rubriques_analyse'] = self.analyze_rubriques(data['lignes_detail'])
                logger.info(f"Rubriques analysées pour {pdf_file.name}: {len(data['rubriques_analyse'])} rubriques trouvées")
            else:
                data['rubriques_analyse'] = []
                logger.info(f"Aucune ligne de détail trouvée pour {pdf_file.name}")
            
            return data
            
        except Exception as e:
            logger.error(f"Erreur lors de l'extraction de {pdf_file.name}: {e}")
            return {
                'nom_fichier': pdf_file.name,
                'erreur': str(e),
                'numero_facture': None,
                'date_facture': None,
                'numero_commande': None,
                'date_echeance': None,
                'destinataire': None,
                'emetteur': 'Select T.T',
                'batch_id': None,
                'assignment_id': None,
                'total_net': None,
                'total_tva': None,
                'total_brut': None,
                'taux_tva': '20%',
                'devise': 'EUR',
                'lignes_detail': []
            }
    
    def extract_invoice_number(self, text: str) -> Optional[str]:
        """Extrait le numéro de facture"""
        patterns = [
            r'Invoice ID/Number[^0-9A-Z]*([0-9]+S[0-9]+)',
            r'(\d{4}S\d{4})',
            r'Invoice.*?(\d{4}S\d{3,4})'
        ]
        
        for pattern in patterns:
            result = self.extract_with_regex(text, pattern)
            if result:
                return result
        return None
    
    def extract_invoice_date(self, text: str) -> Optional[str]:
        """Extrait la date de facture"""
        patterns = [
            r'Invoice Date[^0-9]*(\d{4}/\d{2}/\d{2})',
            r'(\d{4}/\d{2}/\d{2})'
        ]
        
        for pattern in patterns:
            result = self.extract_with_regex(text, pattern)
            if result:
                return result
        return None
    
    def extract_purchase_order(self, text: str) -> Optional[str]:
        """Extrait le numéro de commande"""
        patterns = [
            r'Purchase Order[^0-9]*(\d{10})',
            r'Bon de commande[^0-9]*(\d{10})',
            r'(\d{10})'
        ]
        
        for pattern in patterns:
            result = self.extract_with_regex(text, pattern)
            if result and len(result) == 10:
                return result
        return None
    
    def extract_due_date(self, text: str) -> Optional[str]:
        """Extrait la date d'échéance"""
        patterns = [
            r'Payment Terms[^0-9]*(\d{4}/\d{2}/\d{2})',
            r'Modalités de Paiement[^0-9]*(\d{4}/\d{2}/\d{2})'
        ]
        
        for pattern in patterns:
            result = self.extract_with_regex(text, pattern)
            if result:
                return result
        return None
    
    def extract_recipient(self, text: str) -> str:
        """Extrait le destinataire"""
        if 'Mars Information Services' in text:
            return 'Mars Information Services'
        elif 'Mars Petcare Food France' in text:
            return 'Mars Petcare Food France SAS'
        elif 'Mars' in text:
            return 'Mars (à préciser)'
        return 'Non déterminé'
    
    def extract_batch_id(self, text: str) -> Optional[str]:
        """Extrait le Batch ID"""
        result = self.extract_with_regex(text, r'(\d{4})_\d{5}_')
        return result
    
    def extract_assignment_id(self, text: str) -> Optional[str]:
        """Extrait l'Assignment ID"""
        result = self.extract_with_regex(text, r'\d{4}_(\d{5})_')
        return result
    
    def extract_total_net(self, text: str) -> Optional[float]:
        """Extrait le total net"""
        patterns = [
            r'Invoice Total.*?EUR.*?([\d,\.]+)\s+[\d,\.]+\s+[\d,\.]+',
            r'Invoice Total.*?EUR.*?([\d,\.]+)',
            r'Net Amount.*?([\d,\.]+)',
            r'Montant Net.*?([\d,\.]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return self.normalize_amount(match.group(1))
        return None
    
    def extract_total_vat(self, text: str) -> Optional[float]:
        """Extrait le total TVA"""
        patterns = [
            r'Invoice Total.*?EUR.*?[\d,\.]+\s+([\d,\.]+)\s+[\d,\.]+',
            r'VAT Amount.*?([\d,\.]+)',
            r'Montant TVA.*?([\d,\.]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return self.normalize_amount(match.group(1))
        return None
    
    def extract_total_gross(self, text: str) -> Optional[float]:
        """Extrait le total brut"""
        patterns = [
            r'Invoice Total.*?EUR.*?[\d,\.]+\s+[\d,\.]+\s+([\d,\.]+)',
            r'Gross Amount.*?([\d,\.]+)',
            r'Montant brut.*?([\d,\.]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return self.normalize_amount(match.group(1))
        return None
    
    def extract_line_items(self, text: str) -> List[Dict]:
        """Extrait les lignes de détail avec analyse des rubriques"""
        lines = []
        
        line_patterns = [
            r'(\d{4}_\d{5}_[^0-9]*?)\s+(\d{4}/\d{2}/\d{2})\s+(\w+)\s+([\d,\.]+)\s+(\d+)\s+([\d,\.]+)\s+([\d,\.]+)\s+([\d,\.]+)',
            r'(\d{4}_\d{5}_[^\s]+)\s+(\d{4}/\d{2}/\d{2})\s+(\w+)\s+([\d,\.]+)\s+(\d+)\s+([\d,\.]+)\s+([\d,\.]+)\s+([\d,\.]+)'
        ]
        
        for pattern in line_patterns:
            matches = re.finditer(pattern, text, re.MULTILINE)
            
            for match in matches:
                try:
                    description = match.group(1).strip()
                    batch_id, assignment_id, type_prestation = self.parse_description(description)
                    
                    line_data = {
                        'description': description,
                        'batch_id': batch_id,
                        'assignment_id': assignment_id,
                        'type_prestation': type_prestation,
                        'code_rubrique': self.extract_rubrique_code(description, text),
                        'date_periode': match.group(2),
                        'unite': match.group(3),
                        'prix_unitaire': self.normalize_amount(match.group(4)),
                        'quantite': int(match.group(5)),
                        'montant_net': self.normalize_amount(match.group(6)),
                        'montant_tva': self.normalize_amount(match.group(7)),
                        'montant_brut': self.normalize_amount(match.group(8))
                    }
                    lines.append(line_data)
                except (ValueError, IndexError) as e:
                    logger.warning(f"Erreur lors de l'extraction d'une ligne: {e}")
                    continue
        
        return lines
    
    def parse_description(self, description: str) -> Tuple[str, str, str]:
        """Parse la description pour extraire Batch ID, Assignment ID et type de prestation"""
        try:
            parts = description.split('_')
            if len(parts) >= 3:
                batch_id = parts[0]
                assignment_id = parts[1]
                type_part = '_'.join(parts[2:])
                
                if 'Expense' in type_part:
                    type_prestation = 'Expense'
                elif 'Timesheet' in type_part:
                    type_prestation = 'Timesheet'
                else:
                    type_prestation = 'Autre'
                
                return batch_id, assignment_id, type_prestation
            else:
                return None, None, 'Non déterminé'
        except Exception:
            return None, None, 'Non déterminé'
    
    def extract_rubrique_code(self, description: str, full_text: str) -> Optional[str]:
        """Extrait le code rubrique (ex: OT125) depuis la description ou le texte complet"""
        rubrique_patterns = [
            r'([A-Z]{2}\d{3})',
            r'Code rubrique[^A-Z]*([A-Z]{2}\d{3})',
            r'rubrique[^A-Z]*([A-Z]{2}\d{3})'
        ]
        
        for pattern in rubrique_patterns:
            match = re.search(pattern, description, re.IGNORECASE)
            if match:
                return match.group(1)
        
        for pattern in rubrique_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                return match.group(1)
        
        return None
    
    def analyze_rubriques(self, lines: List[Dict]) -> List[Dict]:
        """Analyse et regroupe les données par rubrique"""
        rubriques_data = {}
        
        for line in lines:
            rubrique_key = f"{line.get('code_rubrique', 'SANS_CODE')}_{line.get('type_prestation', 'SANS_TYPE')}"
            
            if rubrique_key not in rubriques_data:
                rubriques_data[rubrique_key] = {
                    'code_rubrique': line.get('code_rubrique', 'Non déterminé'),
                    'type_prestation': line.get('type_prestation', 'Non déterminé'),
                    'batch_id': line.get('batch_id'),
                    'assignment_id': line.get('assignment_id'),
                    'nb_lignes': 0,
                    'total_quantite': 0,
                    'total_net': 0,
                    'total_tva': 0,
                    'total_brut': 0,
                    'unites': set(),
                    'periodes': set()
                }
            
            rubrique = rubriques_data[rubrique_key]
            rubrique['nb_lignes'] += 1
            rubrique['total_quantite'] += line.get('quantite', 0)
            rubrique['total_net'] += line.get('montant_net', 0) or 0
            rubrique['total_tva'] += line.get('montant_tva', 0) or 0
            rubrique['total_brut'] += line.get('montant_brut', 0) or 0
            
            if line.get('unite'):
                rubrique['unites'].add(line['unite'])
            if line.get('date_periode'):
                rubrique['periodes'].add(line['date_periode'])
        
        for rubrique in rubriques_data.values():
            rubrique['unites'] = ', '.join(sorted(rubrique['unites'])) if rubrique['unites'] else ''
            rubrique['periodes'] = ', '.join(sorted(rubrique['periodes'])) if rubrique['periodes'] else ''
        
        return list(rubriques_data.values())
    
    def process_files(self, uploaded_files) -> List[Dict]:
        """Traite tous les fichiers uploadés"""
        self.extracted_data = []
        
        for uploaded_file in uploaded_files:
            st.write(f"📄 Traitement de: {uploaded_file.name}")
            
            file_content = io.BytesIO(uploaded_file.read())
            file_content.name = uploaded_file.name
            
            data = self.extract_invoice_data(file_content)
            self.extracted_data.append(data)
            
            uploaded_file.seek(0)
        
        return self.extracted_data
    
    def format_worksheet(self, ws, title, has_amounts=False):
        """Applique un formatage professionnel à une feuille"""
        # Couleurs du thème
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Polices
        header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        normal_font = Font(name='Calibri', size=10)
        bold_font = Font(name='Calibri', size=10, bold=True)
        
        # Bordures
        thin_border = Border(
            left=Side(style='thin', color='B3B3B3'),
            right=Side(style='thin', color='B3B3B3'),
            top=Side(style='thin', color='B3B3B3'),
            bottom=Side(style='thin', color='B3B3B3')
        )
        
        # Formater les en-têtes
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Ajuster la hauteur de la ligne d'en-tête
        ws.row_dimensions[1].height = 30
        
        # Formater les lignes de données
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Alterner les couleurs de fond
            if idx % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.start_color.rgb != "FFC000":
                        cell.fill = alt_row_fill
            
            # Appliquer les bordures et l'alignement
            for cell in row:
                cell.border = thin_border
                cell.font = normal_font
                
                # Alignement selon le type de données
                if cell.column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    if has_amounts and cell.column in [col for col in range(1, ws.max_column + 1) 
                                                        if 'EUR' in str(ws.cell(1, col).value) or 
                                                        'Montant' in str(ws.cell(1, col).value) or
                                                        'Total' in str(ws.cell(1, col).value) or
                                                        'Prix' in str(ws.cell(1, col).value)]:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00 €'
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
        
        # Figer la première ligne
        ws.freeze_panes = ws['A2']
    
    def add_summary_sheet(self, writer, workbook):
        """Ajoute une feuille de synthèse au début"""
        # Créer les données de synthèse
        total_factures = len(self.extracted_data)
        factures_ok = sum(1 for d in self.extracted_data if d['numero_facture'])
        total_net = sum(d['total_net'] or 0 for d in self.extracted_data)
        total_tva = sum(d['total_tva'] or 0 for d in self.extracted_data)
        total_brut = sum(d['total_brut'] or 0 for d in self.extracted_data)
        total_lignes = sum(len(d['lignes_detail']) if d['lignes_detail'] else 0 for d in self.extracted_data)
        
        summary_data = {
            'Indicateur': [
                '📊 STATISTIQUES GÉNÉRALES',
                'Nombre de fichiers traités',
                'Extractions réussies',
                'Taux de réussite',
                '',
                '💰 MONTANTS TOTAUX',
                'Total Net (EUR)',
                'Total TVA (EUR)',
                'Total Brut (EUR)',
                '',
                '📋 DÉTAILS',
                'Nombre total de lignes extraites',
                'Moyenne lignes par facture',
                'Date d\'extraction'
            ],
            'Valeur': [
                '',
                total_factures,
                factures_ok,
                f"{(factures_ok/total_factures*100):.1f}%" if total_factures > 0 else '0%',
                '',
                '',
                f"{total_net:,.2f}",
                f"{total_tva:,.2f}",
                f"{total_brut:,.2f}",
                '',
                '',
                total_lignes,
                f"{total_lignes/total_factures:.1f}" if total_factures > 0 else '0',
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='📊 Synthèse', index=False)
        
        ws = writer.sheets['📊 Synthèse']
        
        # Formatage spécial pour la synthèse
        title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        value_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        title_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        section_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        normal_font = Font(name='Calibri', size=10)
        
        # Supprimer les en-têtes par défaut
        ws.delete_rows(1)
        
        # Formater les sections
        section_rows = [1, 6, 11]  # Lignes de titre de section
        
        for idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
            if idx in section_rows:
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[idx].height = 25
            elif row[0].value == '':
                ws.row_dimensions[idx].height = 5
            else:
                row[0].font = normal_font
                row[1].font = Font(name='Calibri', size=10, bold=True)
                row[1].fill = value_fill
                row[0].alignment = Alignment(horizontal='left', vertical='center')
                row[1].alignment = Alignment(horizontal='right', vertical='center')
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
    
    def create_excel_report(self) -> io.BytesIO:
        """Crée un rapport Excel avec les données extraites incluant l'analyse par rubriques"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Feuille 1: Synthèse
            self.add_summary_sheet(writer, writer.book)
            
            # Feuille 2: Résumé des factures
            summary_data = []
            for data in self.extracted_data:
                summary_data.append({
                    'Nom Fichier': data['nom_fichier'],
                    'N° Facture': data['numero_facture'],
                    'Date Facture': data['date_facture'],
                    'N° Commande': data['numero_commande'],
                    'Date Échéance': data['date_echeance'],
                    'Destinataire': data['destinataire'],
                    'Batch ID': data['batch_id'],
                    'Assignment ID': data['assignment_id'],
                    'Total Net (EUR)': data['total_net'],
                    'Total TVA (EUR)': data['total_tva'],
                    'Total Brut (EUR)': data['total_brut'],
                    'Nb Lignes': len(data['lignes_detail']) if data['lignes_detail'] else 0,
                    'Nb Rubriques': len(data['rubriques_analyse']) if data['rubriques_analyse'] else 0,
                    'Statut': '✅ OK' if data['numero_facture'] else '⚠️ Partiel'
                })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='📋 Factures', index=False)
            self.format_worksheet(writer.sheets['📋 Factures'], 'Résumé des Factures', has_amounts=True)
            
            # Feuille 3: Analyse par rubriques
            rubriques_data = []
            for data in self.extracted_data:
                if data.get('rubriques_analyse') and len(data['rubriques_analyse']) > 0:
                    for rubrique in data['rubriques_analyse']:
                        rubriques_data.append({
                            'N° Facture': data['numero_facture'],
                            'N° Commande': data['numero_commande'],
                            'Code Rubrique': rubrique.get('code_rubrique', 'Non déterminé'),
                            'Type Prestation': rubrique.get('type_prestation', 'Non déterminé'),
                            'Batch ID': rubrique.get('batch_id', ''),
                            'Assignment ID': rubrique.get('assignment_id', ''),
                            'Nb Lignes': rubrique.get('nb_lignes', 0),
                            'Quantité': rubrique.get('total_quantite', 0),
                            'Unités': rubrique.get('unites', ''),
                            'Périodes': rubrique.get('periodes', ''),
                            'Net (EUR)': rubrique.get('total_net', 0),
                            'TVA (EUR)': rubrique.get('total_tva', 0),
                            'Brut (EUR)': rubrique.get('total_brut', 0),
                            '% Facture': round((rubrique.get('total_net', 0) / (data['total_net'] or 1)) * 100, 2) if data.get('total_net') and data['total_net'] > 0 else 0
                        })
                else:
                    rubriques_data.append({
                        'N° Facture': data['numero_facture'],
                        'N° Commande': data['numero_commande'],
                        'Code Rubrique': 'TOTAL_FACTURE',
                        'Type Prestation': 'Global',
                        'Batch ID': data.get('batch_id', ''),
                        'Assignment ID': data.get('assignment_id', ''),
                        'Nb Lignes': len(data.get('lignes_detail', [])),
                        'Quantité': 1,
                        'Unités': 'Facture complète',
                        'Périodes': data.get('date_facture', ''),
                        'Net (EUR)': data.get('total_net', 0),
                        'TVA (EUR)': data.get('total_tva', 0),
                        'Brut (EUR)': data.get('total_brut', 0),
                        '% Facture': 100.0
                    })
            
            df_rubriques = pd.DataFrame(rubriques_data)
            df_rubriques.to_excel(writer, sheet_name='🏷️ Rubriques', index=False)
            self.format_worksheet(writer.sheets['🏷️ Rubriques'], 'Analyse par Rubriques', has_amounts=True)
            
            # Feuille 4: Synthèse par type
            synthese_data = {}
            for data in self.extracted_data:
                if data.get('rubriques_analyse') and len(data['rubriques_analyse']) > 0:
                    for rubrique in data['rubriques_analyse']:
                        type_key = rubrique.get('type_prestation', 'Non déterminé')
                        if type_key not in synthese_data:
                            synthese_data[type_key] = {
                                'Type Prestation': type_key,
                                'Nb Factures': 0,
                                'Nb Lignes': 0,
                                'Net (EUR)': 0,
                                'TVA (EUR)': 0,
                                'Brut (EUR)': 0,
                                'Codes Rubriques': set(),
                                'Factures': set()
                            }
                        
                        synthese = synthese_data[type_key]
                        synthese['Nb Lignes'] += rubrique.get('nb_lignes', 0)
                        synthese['Net (EUR)'] += rubrique.get('total_net', 0)
                        synthese['TVA (EUR)'] += rubrique.get('total_tva', 0)
                        synthese['Brut (EUR)'] += rubrique.get('total_brut', 0)
                        synthese['Codes Rubriques'].add(rubrique.get('code_rubrique', 'Non déterminé'))
                        synthese['Factures'].add(data['numero_facture'])
            
            synthese_export = []
            if synthese_data:
                total_net_global = sum(d.get('total_net', 0) for d in self.extracted_data if d.get('total_net'))
                for synthese in synthese_data.values():
                    synthese['Nb Factures'] = len(synthese['Factures'])
                    synthese_export.append({
                        'Type Prestation': synthese['Type Prestation'],
                        'Nb Factures': synthese['Nb Factures'],
                        'Nb Lignes': synthese['Nb Lignes'],
                        'Codes Rubriques': ', '.join(sorted(synthese['Codes Rubriques'])),
                        'Net (EUR)': synthese['Net (EUR)'],
                        'TVA (EUR)': synthese['TVA (EUR)'],
                        'Brut (EUR)': synthese['Brut (EUR)'],
                        '% Total': round((synthese['Net (EUR)'] / total_net_global) * 100, 2) if total_net_global > 0 else 0
                    })
            else:
                total_net_global = sum(d.get('total_net', 0) for d in self.extracted_data if d.get('total_net'))
                total_tva_global = sum(d.get('total_tva', 0) for d in self.extracted_data if d.get('total_tva'))
                total_brut_global = sum(d.get('total_brut', 0) for d in self.extracted_data if d.get('total_brut'))
                
                synthese_export.append({
                    'Type Prestation': 'Extraction globale',
                    'Nb Factures': len(self.extracted_data),
                    'Nb Lignes': sum(len(d.get('lignes_detail', [])) for d in self.extracted_data),
                    'Codes Rubriques': 'Non détectés',
                    'Net (EUR)': total_net_global,
                    'TVA (EUR)': total_tva_global,
                    'Brut (EUR)': total_brut_global,
                    '% Total': 100.0
                })
            
            df_synthese = pd.DataFrame(synthese_export)
            df_synthese.to_excel(writer, sheet_name='📊 Types Prestations', index=False)
            self.format_worksheet(writer.sheets['📊 Types Prestations'], 'Synthèse par Type', has_amounts=True)
            
            # Feuille 5: Détail des lignes
            detail_data = []
            for data in self.extracted_data:
                if data['lignes_detail']:
                    for line in data['lignes_detail']:
                        detail_data.append({
                            'N° Facture': data['numero_facture'],
                            'N° Commande': data['numero_commande'],
                            'Batch ID': line.get('batch_id', ''),
                            'Assignment ID': line.get('assignment_id', ''),
                            'Code Rubrique': line.get('code_rubrique', ''),
                            'Type Prestation': line.get('type_prestation', ''),
                            'Description': line['description'],
                            'Date Période': line['date_periode'],
                            'Unité': line['unite'],
                            'Prix Unit.': line['prix_unitaire'],
                            'Qté': line['quantite'],
                            'Net (EUR)': line['montant_net'],
                            'TVA (EUR)': line['montant_tva'],
                            'Brut (EUR)': line['montant_brut']
                        })
                else:
                    detail_data.append({
                        'N° Facture': data['numero_facture'],
                        'N° Commande': data['numero_commande'],
                        'Batch ID': data['batch_id'],
                        'Assignment ID': data['assignment_id'],
                        'Code Rubrique': '',
                        'Type Prestation': 'Total facture',
                        'Description': 'Total facture',
                        'Date Période': data['date_facture'],
                        'Unité': 'Global',
                        'Prix Unit.': None,
                        'Qté': 1,
                        'Net (EUR)': data['total_net'],
                        'TVA (EUR)': data['total_tva'],
                        'Brut (EUR)': data['total_brut']
                    })
            
            df_detail = pd.DataFrame(detail_data)
            df_detail.to_excel(writer, sheet_name='📝 Détail Lignes', index=False)
            self.format_worksheet(writer.sheets['📝 Détail Lignes'], 'Détail des Lignes', has_amounts=True)
        
        output.seek(0)
        return output


def main():
    st.title("📄 PDF Extractor Select T.T")
    st.markdown("### Application d'extraction automatique des auto-factures Select T.T")
    
    # Sidebar
    st.sidebar.header("📋 Instructions")
    st.sidebar.markdown("""
    1. **Uploadez** vos fichiers PDF
    2. **Lancez** l'extraction
    3. **Vérifiez** les résultats
    4. **Téléchargez** le fichier Excel
    """)
    
    st.sidebar.header("📊 Données extraites")
    st.sidebar.markdown("""
    - N° Auto-facture
    - N° Commande  
    - Dates (facture, échéance)
    - Destinataire
    - Batch ID & Assignment ID
    - Montants (Net, TVA, Brut)
    - Détail des lignes
    - Analyse par rubriques
    """)
    
    # Upload des fichiers
    st.header("📤 Upload des fichiers PDF")
    uploaded_files = st.file_uploader(
        "Sélectionnez vos fichiers PDF d'auto-factures",
        type=['pdf'],
        accept_multiple_files=True,
        help="Vous pouvez sélectionner plusieurs fichiers PDF à la fois"
    )
    
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} fichier(s) sélectionné(s)")
        
        # Afficher la liste des fichiers
        with st.expander("📁 Fichiers sélectionnés", expanded=True):
            for i, file in enumerate(uploaded_files, 1):
                st.write(f"{i}. {file.name} ({file.size / 1024:.1f} KB)")
        
        # Bouton d'extraction
        if st.button("🚀 Lancer l'extraction", type="primary"):
            with st.spinner("Extraction en cours..."):
                extractor = PDFExtractor()
                
                # Barre de progression
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    extracted_data = extractor.process_files(uploaded_files)
                    progress_bar.progress(1.0)
                    status_text.text("✅ Extraction terminée!")
                    
                    # Affichage des résultats
                    st.header("📊 Résultats de l'extraction")
                    
                    # Statistiques enrichies
                    col1, col2, col3, col4, col5 = st.columns(5)
                    
                    with col1:
                        st.metric("📄 Fichiers traités", len(extracted_data))
                    
                    with col2:
                        success_count = sum(1 for d in extracted_data if d['numero_facture'])
                        st.metric("✅ Extractions réussies", success_count)
                    
                    with col3:
                        total_net = sum(d['total_net'] or 0 for d in extracted_data)
                        st.metric("💰 Total Net (EUR)", f"{total_net:,.2f}")
                    
                    with col4:
                        total_gross = sum(d['total_brut'] or 0 for d in extracted_data)
                        st.metric("💰 Total Brut (EUR)", f"{total_gross:,.2f}")
                    
                    with col5:
                        total_lignes = sum(len(d['lignes_detail']) if d['lignes_detail'] else 0 for d in extracted_data)
                        st.metric("📋 Lignes extraites", total_lignes)
                    
                    # Nouvelle section : Analyse par rubriques
                    st.subheader("🏷️ Analyse par rubriques")
                    
                    # Regrouper toutes les rubriques
                    all_rubriques = []
                    for data in extracted_data:
                        if data['rubriques_analyse']:
                            for rubrique in data['rubriques_analyse']:
                                rubrique['numero_facture'] = data['numero_facture']
                                all_rubriques.append(rubrique)
                    
                    if all_rubriques:
                        # Créer un tableau synthétique par type de prestation
                        types_prestations = {}
                        for rubrique in all_rubriques:
                            type_key = rubrique['type_prestation']
                            if type_key not in types_prestations:
                                types_prestations[type_key] = {
                                    'nb_factures': 0,
                                    'nb_lignes': 0,
                                    'total_net': 0,
                                    'total_brut': 0,
                                    'factures': set()
                                }
                            
                            types_prestations[type_key]['nb_lignes'] += rubrique['nb_lignes']
                            types_prestations[type_key]['total_net'] += rubrique['total_net']
                            types_prestations[type_key]['total_brut'] += rubrique['total_brut']
                            types_prestations[type_key]['factures'].add(rubrique['numero_facture'])
                        
                        # Affichage des métriques par type
                        st.markdown("**Répartition par type de prestation:**")
                        cols = st.columns(len(types_prestations))
                        
                        for i, (type_name, stats) in enumerate(types_prestations.items()):
                            with cols[i]:
                                st.metric(
                                    f"🔧 {type_name}",
                                    f"{stats['total_net']:,.2f} €",
                                    f"{stats['nb_lignes']} lignes"
                                )
                        
                        # Tableau détaillé des rubriques
                        rubriques_display = []
                        for rubrique in all_rubriques:
                            rubriques_display.append({
                                'Facture': rubrique['numero_facture'],
                                'Code Rubrique': rubrique['code_rubrique'] or '❌',
                                'Type': rubrique['type_prestation'],
                                'Lignes': rubrique['nb_lignes'],
                                'Quantité': rubrique['total_quantite'],
                                'Net (€)': f"{rubrique['total_net']:,.2f}",
                                'Brut (€)': f"{rubrique['total_brut']:,.2f}"
                            })
                        
                        if rubriques_display:
                            st.dataframe(pd.DataFrame(rubriques_display), use_container_width=True)
                    else:
                        st.info("ℹ️ Aucune rubrique détaillée trouvée dans les PDFs")
                    
                    # Tableau de résultats principal
                    st.subheader("📋 Détail des extractions")
                    
                    display_data = []
                    for data in extracted_data:
                        display_data.append({
                            'Fichier': data['nom_fichier'],
                            'N° Facture': data['numero_facture'] or '❌',
                            'N° Commande': data['numero_commande'] or '❌',
                            'Date': data['date_facture'] or '❌',
                            'Destinataire': data['destinataire'],
                            'Total Net': f"{data['total_net']:,.2f} €" if data['total_net'] else '❌',
                            'Total Brut': f"{data['total_brut']:,.2f} €" if data['total_brut'] else '❌',
                            'Statut': '✅' if data['numero_facture'] else '⚠️'
                        })
                    
                    df_display = pd.DataFrame(display_data)
                    st.dataframe(df_display, use_container_width=True)
                    
                    # Génération du fichier Excel
                    st.header("💾 Export Excel")
                    
                    with st.spinner("Génération du fichier Excel avec formatage professionnel..."):
                        excel_file = extractor.create_excel_report()
                    
                    # Bouton de téléchargement
                    filename = f"Extraction_PDF_SelectTT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    
                    st.download_button(
                        label="📊 Télécharger le fichier Excel",
                        data=excel_file.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
                    st.success("🎉 Extraction terminée ! Vous pouvez télécharger le fichier Excel formaté.")
                    
                    # Informations sur le fichier Excel
                    with st.expander("ℹ️ Contenu du fichier Excel"):
                        st.markdown("""
                        **Le fichier Excel contient 5 feuilles formatées professionnellement :**
                        
                        1. **📊 Synthèse** : Vue d'ensemble avec statistiques globales
                        2. **📋 Factures** : Résumé de toutes les factures avec indicateurs
                        3. **🏷️ Rubriques** : Analyse détaillée par rubrique et type de prestation
                        4. **📊 Types Prestations** : Synthèse consolidée par type de prestation
                        5. **📝 Détail Lignes** : Détail complet ligne par ligne
                        
                        **Fonctionnalités du fichier :**
                        - ✨ Formatage professionnel avec couleurs et styles
                        - 📏 Colonnes auto-ajustées pour une lecture optimale
                        - 🔢 Montants formatés en euros avec séparateurs de milliers
                        - 🎨 Lignes alternées pour faciliter la lecture
                        - ❄️ En-têtes figés pour navigation facile
                        - 📊 Données prêtes pour analyse et tableaux croisés dynamiques
                        """)
                
                except Exception as e:
                    st.error(f"❌ Erreur lors de l'extraction: {e}")
                    logger.error(f"Erreur: {e}")
    
    else:
        st.info("👆 Commencez par uploader vos fichiers PDF d'auto-factures")
    
    # Footer
    st.markdown("---")
    st.markdown("**PDF Extractor Select T.T** - Version 2.0 | Excel avec formatage professionnel")


if __name__ == "__main__":
    main()
